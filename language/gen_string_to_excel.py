# -*- coding: utf-8 -*-

# 1.0.0
# WPR
# 20210714

import sys
import os
import openpyxl

def main(argv):
    gCurrDir = os.getcwd()
    gLanguageFilePath = gCurrDir + "/language_table.c"
    gExcelFilePath = gCurrDir + "/language-gen.xlsx"

    gExcelWorkbook = openpyxl.Workbook()
    gAllDataSheet = gExcelWorkbook.create_sheet("language", index=0)
    gAllDataRow = 1

    with open(gLanguageFilePath, 'rb') as languageFileHand:
        
        isStart = False
        isHasTitle = False

        orgOneLineData = languageFileHand.readline()
        while orgOneLineData:
            orgOneLineString = orgOneLineData.decode().strip()

            if(not isStart):
                if("language_strings_struct" in orgOneLineString):
                    isStart = True
            else:
                if(orgOneLineString.startswith("};")):
                    break

                if(orgOneLineString.startswith("//")):
                    if(not isHasTitle):
                        titleString = orgOneLineString.replace("//", '').replace("，", ',').strip()
                        titleList = titleString.split(',')

                        for rowIndex in range(len(titleList)):
                            gAllDataSheet.cell(gAllDataRow, rowIndex+1+1, titleList[rowIndex].strip())

                        isHasTitle = True
                        gAllDataRow += 1
                        print(titleList)

                elif (orgOneLineString.startswith("{")):

                    if("STRING_" in orgOneLineString):
                        contentString = orgOneLineString.replace("{", '').replace("},", '').strip()
                        keyIndex = contentString.index(",")
                        KeyString = contentString[:keyIndex]

                        gAllDataSheet.cell(gAllDataRow, 1, KeyString.strip())
                        
                        textString = contentString[keyIndex+1:]
                        contentList = textString.split('",')
                        for rowIndex in range(len(contentList)):
                            itemString = contentList[rowIndex].replace('"', '').strip()
                            gAllDataSheet.cell(gAllDataRow, rowIndex+1+1, itemString)
                        gAllDataRow += 1
                    elif (len(orgOneLineString) > 1):
                        print("Error string:"+orgOneLineString)

            orgOneLineData = languageFileHand.readline()

    gExcelWorkbook.save(gExcelFilePath)
    gExcelWorkbook.close()
    

if __name__ == "__main__":
    main(sys.argv)
    print("OK")
